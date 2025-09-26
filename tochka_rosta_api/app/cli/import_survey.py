"""CLI tool to import survey definitions from Excel workbooks."""

from __future__ import annotations

import argparse
import asyncio
import json
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..core.database import async_session_factory
from ..models.survey import SurveyOption, SurveyQuestion, SurveySection, SurveyVersion


@dataclass(slots=True)
class SectionRow:
    code: str
    title: str
    order: int
    source_row: int


@dataclass(slots=True)
class QuestionRow:
    section_code: str
    code: str
    text: str
    question_type: str
    order: int
    required: bool
    meta: dict[str, Any] | None
    source_row: int


@dataclass(slots=True)
class OptionRow:
    question_code: str
    value: str
    label: str
    order: int
    source_row: int


class ErrorCollector:
    def __init__(self) -> None:
        self._errors: list[str] = []

    def add(self, sheet: str, row: int | None, message: str) -> None:
        if row is None:
            self._errors.append(f"[{sheet}] {message}")
        else:
            self._errors.append(f"[{sheet}] Row {row}: {message}")

    @property
    def has_errors(self) -> bool:
        return bool(self._errors)

    def report(self) -> str:
        return "\n".join(self._errors)


def _normalize_header(cell_value: Any) -> str:
    return str(cell_value).strip().lower() if cell_value is not None else ""


def _row_value(row: tuple[Any, ...], index: int) -> Any:
    if index < 0:
        return None
    try:
        return row[index]
    except IndexError:
        return None


def _extract_headers(sheet, errors: ErrorCollector) -> dict[str, int]:
    try:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        errors.add(sheet.title, None, "Sheet is empty")
        return {}
    header_map: dict[str, int] = {}
    for index, value in enumerate(first_row):
        key = _normalize_header(value)
        if key:
            header_map[key] = index
    return header_map


def parse_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    value_str = str(value).strip().lower()
    return value_str in {"1", "true", "yes", "y", "да", "истина"}


def parse_sections(workbook, errors: ErrorCollector) -> list[SectionRow]:
    if "Sections" not in workbook:
        errors.add("Sections", None, "Sheet not found")
        return []

    sheet = workbook["Sections"]
    header_map = _extract_headers(sheet, errors)
    required = {"code", "title", "order"}
    missing = required - set(header_map)
    if missing:
        errors.add("Sections", None, f"Missing required columns: {', '.join(sorted(missing))}")
        return []
    rows: list[SectionRow] = []
    seen_codes: set[str] = set()

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        code_raw = _row_value(row, header_map["code"])
        title_raw = _row_value(row, header_map["title"])
        order_raw = _row_value(row, header_map["order"])

        if not code_raw or not str(code_raw).strip():
            errors.add("Sections", index, "Missing section code")
            continue
        code = str(code_raw).strip()

        if code in seen_codes:
            errors.add("Sections", index, f"Duplicate section code '{code}'")
            continue
        seen_codes.add(code)

        if not title_raw or not str(title_raw).strip():
            errors.add("Sections", index, "Missing section title")
            continue
        title = str(title_raw).strip()

        try:
            order = int(order_raw)
        except (TypeError, ValueError):
            errors.add("Sections", index, "Order must be an integer")
            continue

        rows.append(SectionRow(code=code, title=title, order=order, source_row=index))

    return rows


def parse_questions(workbook, errors: ErrorCollector) -> list[QuestionRow]:
    if "Questions" not in workbook:
        errors.add("Questions", None, "Sheet not found")
        return []

    sheet = workbook["Questions"]
    header_map = _extract_headers(sheet, errors)
    required_columns = {"section_code", "code", "text", "type", "required"}
    missing = required_columns - set(header_map)
    if missing:
        errors.add("Questions", None, f"Missing required columns: {', '.join(sorted(missing))}")
        return []
    rows: list[QuestionRow] = []
    seen_codes: set[str] = set()
    auto_order: dict[str, int] = defaultdict(int)

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        def get_value(column: str) -> Any:
            position = header_map.get(column)
            if position is None:
                return None
            return _row_value(row, position)

        section_code_raw = get_value("section_code")
        code_raw = get_value("code")
        text_raw = get_value("text")
        type_raw = get_value("type")
        required_raw = get_value("required")
        meta_raw = get_value("meta")
        order_raw = get_value("order")

        if not section_code_raw or not str(section_code_raw).strip():
            errors.add("Questions", index, "Missing section_code")
            continue
        section_code = str(section_code_raw).strip()

        if not code_raw or not str(code_raw).strip():
            errors.add("Questions", index, "Missing question code")
            continue
        code = str(code_raw).strip()
        if code in seen_codes:
            errors.add("Questions", index, f"Duplicate question code '{code}'")
            continue
        seen_codes.add(code)

        if not text_raw or not str(text_raw).strip():
            errors.add("Questions", index, "Missing question text")
            continue
        text = str(text_raw).strip()

        if not type_raw or not str(type_raw).strip():
            errors.add("Questions", index, "Missing question type")
            continue
        question_type = str(type_raw).strip()

        required = parse_bool(required_raw)

        meta: dict[str, Any] | None = None
        if meta_raw not in (None, ""):
            meta_text = str(meta_raw).strip()
            try:
                parsed_meta = json.loads(meta_text)
            except json.JSONDecodeError as exc:
                errors.add("Questions", index, f"Invalid meta JSON: {exc.msg}")
                continue
            if not isinstance(parsed_meta, dict):
                errors.add("Questions", index, "Meta must be a JSON object")
                continue
            meta = parsed_meta

        try:
            order = int(order_raw)
        except (TypeError, ValueError):
            auto_order[section_code] += 1
            order = auto_order[section_code]

        rows.append(
            QuestionRow(
                section_code=section_code,
                code=code,
                text=text,
                question_type=question_type,
                order=order,
                required=required,
                meta=meta,
                source_row=index,
            )
        )

    return rows


def parse_options(workbook, errors: ErrorCollector) -> list[OptionRow]:
    if "Options" not in workbook:
        errors.add("Options", None, "Sheet not found")
        return []

    sheet = workbook["Options"]
    header_map = _extract_headers(sheet, errors)
    required = {"question_code", "value", "label", "order"}
    missing = required - set(header_map)
    if missing:
        errors.add("Options", None, f"Missing required columns: {', '.join(sorted(missing))}")
        return []
    rows: list[OptionRow] = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        def get_value(column: str) -> Any:
            position = header_map.get(column)
            if position is None:
                return None
            return _row_value(row, position)

        question_code_raw = get_value("question_code")
        value_raw = get_value("value")
        label_raw = get_value("label")
        order_raw = get_value("order")

        if not question_code_raw or not str(question_code_raw).strip():
            errors.add("Options", index, "Missing question_code")
            continue
        question_code = str(question_code_raw).strip()

        if value_raw is None or str(value_raw).strip() == "":
            errors.add("Options", index, "Missing option value")
            continue
        value = str(value_raw).strip()

        if label_raw is None or str(label_raw).strip() == "":
            errors.add("Options", index, "Missing option label")
            continue
        label = str(label_raw).strip()

        try:
            order = int(order_raw)
        except (TypeError, ValueError):
            errors.add("Options", index, "Order must be an integer")
            continue

        rows.append(
            OptionRow(
                question_code=question_code,
                value=value,
                label=label,
                order=order,
                source_row=index,
            )
        )

    return rows


async def persist_to_database(
    *,
    title: str,
    activate: bool,
    sections: list[SectionRow],
    questions: list[QuestionRow],
    options: list[OptionRow],
) -> SurveyVersion:
    async with async_session_factory() as session:
        version = SurveyVersion(title=title, is_active=activate)
        session.add(version)
        await session.flush()

        section_map: dict[str, SurveySection] = {}
        for section_row in sections:
            section = SurveySection(
                version_id=version.id,
                code=section_row.code,
                title=section_row.title,
                order=section_row.order,
            )
            session.add(section)
            await session.flush()
            section_map[section_row.code] = section

        question_map: dict[str, SurveyQuestion] = {}
        for question_row in questions:
            section = section_map[question_row.section_code]
            question = SurveyQuestion(
                section_id=section.id,
                code=question_row.code,
                text=question_row.text,
                question_type=question_row.question_type,
                order=question_row.order,
                required=question_row.required,
                meta=question_row.meta,
            )
            session.add(question)
            await session.flush()
            question_map[question_row.code] = question

        for option_row in options:
            question = question_map.get(option_row.question_code)
            if question is None:
                # This should not happen if validation passed but guard just in case.
                raise ValueError(f"Unknown question code '{option_row.question_code}' for option")
            option = SurveyOption(
                question_id=question.id,
                value=option_row.value,
                label=option_row.label,
                order=option_row.order,
            )
            session.add(option)

        await session.commit()
        await session.refresh(version)
        return version


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the Excel workbook (.xlsx)")
    parser.add_argument("--title", required=True, help="Title for the new survey version")
    parser.add_argument(
        "--activate",
        action="store_true",
        default=False,
        help="Activate the imported survey version (default: inactive)",
    )
    return parser


def validate_references(
    *,
    sections: list[SectionRow],
    questions: list[QuestionRow],
    options: list[OptionRow],
    errors: ErrorCollector,
) -> None:
    section_codes = {section.code for section in sections}
    for question in questions:
        if question.section_code not in section_codes:
            errors.add("Questions", question.source_row, f"Unknown section_code '{question.section_code}'")

    question_codes = {question.code for question in questions}
    for option in options:
        if option.question_code not in question_codes:
            errors.add("Options", option.source_row, f"Unknown question_code '{option.question_code}'")


async def main_async(args: argparse.Namespace) -> bool:
    workbook_path: Path = args.workbook
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook '{workbook_path}' does not exist")

    workbook = load_workbook(filename=workbook_path, data_only=True)
    errors = ErrorCollector()

    sections = parse_sections(workbook, errors)
    questions = parse_questions(workbook, errors)
    options = parse_options(workbook, errors)

    validate_references(sections=sections, questions=questions, options=options, errors=errors)

    if errors.has_errors:
        print("Import failed due to the following errors:")
        print(errors.report())
        return False

    version = await persist_to_database(
        title=args.title,
        activate=args.activate,
        sections=sections,
        questions=questions,
        options=options,
    )

    print("Import completed successfully.")
    print(f"Survey version ID: {version.id}")
    print(f"Sections imported: {len(sections)}")
    print(f"Questions imported: {len(questions)}")
    print(f"Options imported: {len(options)}")
    return True


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()
    success = asyncio.run(main_async(args))
    if not success:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
